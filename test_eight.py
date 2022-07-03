import openpyxl


def test_excel_data_read(setup):
    wb=openpyxl.load_workbook("C:\\Users\\Nikita\\PycharmProjects\\pytestDemo\\test_data.xlsx")
    sheet=wb.active  #wb.get_sheet_by_name("sheet1")

    rows=sheet.max_row
    cols=sheet.max_column

    for r in range(1,rows+1):
     for c in range(1,cols+1):
         if sheet.cell(row=r,column=c).value is None:
             sheet.cell(row=r, column=c).value=r
             print(sheet.cell(row=r, column=c).value, end=" ")
             break
     else:
         print(sheet.cell(row=r,column=c).value,end=" ")
     print(" ")

     wb.save("C:\\Users\\Nikita\\PycharmProjects\\pytestDemo\\test_data.xlsx")
